import streamlit as st
import pandas as pd
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

if 'globalList' not in st.session_state:
    st.session_state['globalList'] = []

if 'file_name' not in st.session_state:
    st.session_state['file_name'] = ""

if 'input_data' not in st.session_state:
    st.session_state['input_data'] = ""

if 'header_data' not in st.session_state:
    st.session_state['header_data'] = ""


# find out what the index should be in a giving dataframe
# Return the index of the row that should be the header
def findHeader(df):
    old_header = df.columns.tolist()
    none_empty_cell = 0
    for each in old_header:
        if "Unnamed" not in each:
            none_empty_cell += 1
    print(none_empty_cell)
    newList = []
    for i in range(0,len(df.index)-1):
        counter = 0
        for cell in df.iloc[i]:
            if pd.isnull(cell):
                counter += 1
        newList.append(counter)
    minVal = min(newList)
    firstindex = newList.index(minVal)
    lastindex = len(newList) - newList[::-1].index(minVal)
    if len(df.iloc[firstindex]) == none_empty_cell:
        df.columns = old_header
        df = df[firstindex:lastindex]
    else:
        df.columns = df.iloc[firstindex]
        df = df[firstindex+1:lastindex]
    return df

def GRIdf(df,GRI):
    for i in range(0, len(df.index) - 1):
        for index, cell in enumerate(df.iloc[i]):
            if type(cell) is not str:
                df.iat[i, index] = round(float(cell) * (1+GRI/100), 2)
    return df

st.set_page_config(layout="wide")
st.title("Rate Table Generator")
uploaded_file = st.file_uploader("Choose a file")
current, proposed, accessorial = st.tabs(['Current','Proposed','Header & Accessorials'])
with st.sidebar:
    st.title("Options Manager")

    if uploaded_file is not None:
        mode = st.radio("GRI Apply To All", [True, "With Condition"])
        if mode == True:
            GRI = st.number_input("GRI %", step=0.01)
            st.session_state['GRI'] = GRI
        if mode == "With Condition":
            st.warning("Rules must be mutually exclusive")
            data = pd.read_excel(uploaded_file)
            df = findHeader(data)
            columnName = st.selectbox("Filter On Column:", list(df))
            value = st.multiselect("Equal One Or Multiple Conditions", df[columnName].unique())
            GRIvalue = st.text_input("GRI Value - Percentage")
            if st.button("Add Rule"):
                st.session_state['globalList'].append({columnName:value,"GRI":int(GRIvalue)})
            if st.button("Clear Rules"):
                st.session_state['globalList'] =[]
            st.write(st.session_state['globalList'])
            print(st.session_state['globalList'])
if uploaded_file is not None:
    with current:
        data = pd.read_excel(uploaded_file)
        df = findHeader(data)
        st.dataframe(df, use_container_width=True)

    with accessorial:
        agree = st.checkbox("Include Standard Headers & Accessorials")
        if agree:

            allow_editing = st.radio("Edit Terms", [False, True])

            st.subheader("Headers")
            header_data = f"""		
            CCLS Rates - CFF		
            For Customer - KV		
            Effective Start Date: 2022-05-01		
            Shipping Location: Brampton, ON		
            """

            if allow_editing:
                output_header = st.text_area("Terms", value=header_data, height=200)
                st.session_state['header_data'] = output_header
            else:
                st.text(header_data)

            input_data ="""
            Rates as outlined, are valid for thirty (30) days after issuance and subject to acceptance\n
            Rates are presented in Canadian Dollars
            Rates are subject to weekly posted FCA fuel surcharge in effect at the time of shipping
            Rates as outlined apply to freight that can be easily and safely conveyed by forklift or pallet truck
            Except for FTL, rates are applicable for $ per 100 pounds
            Rates are applied in pounds
            Weights are rounded up to the next pound
            Rates are subject to a 10 pound density
            Rates are subject to change with 30 day notice
            Shipments weighing 10,000 Lbs. or greater are subject to FTL fuel surcharge
            Linear Foot:  Shipments occupying 10 Ft or more of trailer space will be calculated at 1,000 Lbs. per ft.
            Non-Stack:  Non-Stack:  Any pallet or shipping unit that is deemed as non-stackable may be subject to a height rule of 96”
            Liability:  Calculated at $2/Lb. computed on total weight of shipment unless a higher value has been declared
            Shipments are subject to applicable beyond and/or accessorial surcharges processed at CCLS cost +15%
            Services are subject to carrier standard Terms and Conditions
            """

            st.subheader("Standard Terms")
            if allow_editing:
                output_term = st.text_area("Terms", value=input_data, height=400)
                st.session_state['input_data'] = output_term
            else:
                st.text(input_data)

    with proposed:
        if mode == True:
            if GRI == 0:
                st.warning("Please select GRI %")
            st.success(f'A GRI of {GRI}% have been applied.')
            downloaddf = GRIdf(df, GRI)
            file_name = st.text_input(
                "Please Enter Name of the File 👇",
            )
            st.session_state["file_name"] = file_name
            if st.button("Save Data"):
                with st.spinner(text="In progress..."):
                    downloaddf.to_excel("Rate Table.xlsx", index = False, header = True, sheet_name="rate")
                    if agree:
                        wb2 = load_workbook("Rate Table.xlsx")
                        wb2.create_sheet("additional info")
                        wb2['additional info'].cell(1,1).value = st.session_state['header_data']
                        wb2['additional info'].cell(1,1).alignment=Alignment(wrap_text=True)
                        wb2['additional info'].cell(2,1).value = st.session_state['input_data']
                        wb2['additional info'].cell(2,1).alignment=Alignment(wrap_text=True)
                        wb2.save("Rate Table.xlsx")

                    with open("Rate Table.xlsx", 'rb') as my_file:
                        st.download_button(label='Download', data=my_file, file_name=f'{st.session_state["file_name"]}.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            st.dataframe(downloaddf, use_container_width=True)
        elif mode == "With Condition":
            st.write("Current Parsing Condition:")
            df = findHeader(data)
            dfList = []
            if not st.session_state['globalList']:
                st.warning("Please add condition for filter:")
            else:
                st.write(st.session_state['globalList'])
                for each in st.session_state['globalList']:
                    columnName = next(iter(each))
                    dfpiece = df.loc[df[columnName].isin(each[columnName])]
                    dfpiece = GRIdf(dfpiece,each['GRI'])
                    dfList.append(dfpiece)
                finaldf = pd.concat(dfList)

                if st.button("Save Data"):
                    with st.spinner(text="In progress..."):
                        finaldf.to_excel("Rate Table.xlsx", index=False, header=True, sheet_name="rate")
                        if agree:
                            wb2 = load_workbook("Rate Table.xlsx")
                            wb2.create_sheet("additional info")
                            wb2['additional info'].cell(1, 1).value = st.session_state['header_data']
                            wb2['additional info'].cell(row=15, column=1).value = st.session_state['input_data']
                            wb2.save("Rate Table.xlsx")
                        with open("Rate Table.xlsx", 'rb') as my_file:
                            st.download_button(label='Download', data=my_file, file_name='Rate Table.xlsx',
                                               mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                st.dataframe(finaldf, use_container_width=True)